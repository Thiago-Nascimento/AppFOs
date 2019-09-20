from tkinter import * 
from tkinter import messagebox, ttk
from openpyxl import *
import datetime, os, sys, getpass

root = Tk()
root.title("Analisa FO v1.2.0")

root.state('zoomed')

# Changing to the directory
os.chdir(r"\\polarsu2\Geqma\Engenharia da Qualidade\Laudos\FO\APP\BASE DADOS")

emptyRow = None
definedRow = False

foNumber = ""
arriveDate = ""
arriveHour = ""
deliveringDate = ""
deliveringHour = ""
defect = ""
testType = ""
foTest = ""
numberOfTestingPoints = 0
register = ""
comment = ""
idString = ""

analisedFos = []

wb = load_workbook("results.xlsx")
resultSheet = wb["resultsSheet"]

for b in range(2, 1048576):
	e = resultSheet.cell(row=b, column=5)

	if e.value != None and e.value != "":
		analisedFos.append(e.value)
	else:
		break	

wb.close()

print("Analised FOs")

registeredFos = []
comboBoxTextValues = []

foNumberValues = []
arriveDateValues = []
arriveHourValues = []
deliveringDateValues = []
deliveringHourValues = []
defectValues = []
testTypeValues = []
valuesList = []
numberOfTestingPointsValues = []
registerValues = []
commentValues = []
idValues = []

registeredFoNumberValues = []
registeredArriveDateValues = []
registeredArriveHourValues = []
registeredDeliveringDateValues = []
registeredDeliveringHourValues = []
registeredDefectValues = []
registeredTestTypeValues = []
registeredValuesList = []
registeredNumberOfTestingPointsValues = []
registeredRegisterValues = []
registeredCommentValues = []
registeredIdValues = []

testingPointsLabelList = {}
testingPointsEntryList = {}

resultList = {}

analised = []

wb = load_workbook("db.xlsx")
ws = wb["fosParaAnalise"]

for z in range(2,1048576):
	# 1048576 is the max number of rows in excel
	a = ws.cell(row=z, column=1)
	b = ws.cell(row=z, column=2)
	c = ws.cell(row=z, column=3)
	d = ws.cell(row=z, column=4)
	e = ws.cell(row=z, column=5)
	f = ws.cell(row=z, column=6)
	g = ws.cell(row=z, column=7)
	h = ws.cell(row=z, column=8)
	I = ws.cell(row=z, column=9)
	j = ws.cell(row=z, column=10)
	k = ws.cell(row=z, column=11)
	l = ws.cell(row=z, column=12)
	
	if l.value != None and l.value != "":

		if l.value not in analisedFos:
		
			registeredFos.append(a.value)
			comboBoxTextValues.append(a.value)

			registeredFoNumberValues.append(a.value)
			foNumberValues.append(a.value)

			registeredArriveDateValues.append(b.value)
			arriveDateValues.append(b.value)
			
			registeredArriveHourValues.append(c.value)
			arriveHourValues.append(c.value)

			registeredDeliveringDateValues.append(d.value)
			deliveringDateValues.append(d.value)

			registeredDeliveringHourValues.append(e.value)
			deliveringHourValues.append(e.value)

			registeredDefectValues.append(f.value)
			defectValues.append(f.value)

			registeredTestTypeValues.append(g.value)
			testTypeValues.append(g.value)

			registeredNumberOfTestingPointsValues.append(h.value)
			numberOfTestingPointsValues.append(h.value)
			
			registeredRegisterValues.append(I.value)
			registerValues.append(I.value)

			registeredCommentValues.append(j.value)
			commentValues.append(j.value)

			registeredValuesList.append(k.value)
			valuesList.append(k.value)

			registeredIdValues.append(l.value)
			idValues.append(l.value)

	elif l.value == None or l.value == "":
		# print("Lists Built...")
		break

wb.close()

# print(f"valuesList: {idValues}")
print(f"valuesList length: {len(idValues)}")

# print("Final List:", valuesList)

# print("Tests:", testTypeValues)

# close the db.xlsx, and open the results.xlsx
wb.close()

wb = load_workbook("results.xlsx")
resultSheet = wb["resultsSheet"]

def saveTestResults():
	global foNumber
	global defect
	global testType
	global resultList
	global definedRow
	global emptyRow

	if testType != "000-TESTES DE COR":

		for index in range(numberOfTestingPoints):
			resultList[index] = testingPointsEntryList[index].get()
		
		if analistNameEntry.get() != "":
		
			if testingPointsEntryList[index].get() != "" and testingPointsEntryList[0].get() != "":		
				if definedRow == False:
					for i in range(2,1048576):
						# 1048576 is the max number of rows in excel
						c = resultSheet.cell(row=i, column=1)
						
						if c.value == None or c.value == "":
							emptyRow = i
							definedRow = True
							break

				# print('Got Empty Row...')
				# print('Empty Row:',emptyRow)

				date = str(datetime.datetime.now().date().strftime("%d-%m-%Y"))
				time = str(datetime.datetime.now().time())	

				resultSheet.cell(row=emptyRow, column=1).value = foNumber	
				resultSheet.cell(row=emptyRow, column=2).value = defect
				resultSheet.cell(row=emptyRow, column=3).value = register[:16]
				resultSheet.cell(row=emptyRow, column=4).value = comment
				resultSheet.cell(row=emptyRow, column=5).value = idString
				resultSheet.cell(row=emptyRow, column=6).value = analistNameEntry.get()
				resultSheet.cell(row=emptyRow, column=7).value = str(getpass.getuser())
				resultSheet.cell(row=emptyRow, column=8).value = date + " " + time[:8]
				resultSheet.cell(row=emptyRow, column=9).value = testType
				resultSheet.cell(row=emptyRow, column=10).value = foTest
				resultSheet.cell(row=emptyRow, column=11).value = numberOfTestingPoints

				for index in range(numberOfTestingPoints):
					columnPosition = index + 12
					resultSheet.cell(row=emptyRow, column=columnPosition).value = resultList[index]

				# print(resultList)				
				
				try:
					wb.save("results.xlsx")
					messagebox.showinfo("Sucesso", "Resultados cadastrados com sucesso.")

					# clearEntries
					for i in range(numberOfTestingPoints):
						testingPointsEntryList[i].delete(0, END)				

					analised.append(idString)
					
					other_index	= idValues.index(idString)

					idValues.remove(idString)

					del foNumberValues[other_index]
					del arriveDateValues[other_index]
					del arriveHourValues[other_index]
					del deliveringDateValues[other_index]
					del deliveringHourValues[other_index]
					del defectValues[other_index]
					del testTypeValues[other_index]
					del numberOfTestingPointsValues[other_index]
					del registerValues[other_index]
					del commentValues[other_index]
					del valuesList[other_index]
					
					definedRow = False

					print(f"Saved: valuesList length: {len(valuesList)}")
					# print(idValues)

					goBack()	

				except Exception as e:
					print(e)
					messagebox.showerror("Erro", "Não foi possível fazer o cadastro, a planilha está aberta em outro lugar.\nFeche-a e tente novamente.")

			else:
				# print("There is a test missing...")
				messagebox.showerror("Erro", "Preencha todos os campos.")	

		else:
			messagebox.showerror("Erro", "Digite seu Nome.")
	
	elif testType == "000-TESTES DE COR":

		if analistNameEntry.get() != "":

			entriesList = [entryLuminTela.get(),
				entryCoordATela.get(),
				entryCoordBTela.get(),
				entryAlvuraISOTela.get(),
				entryBrancuraTela.get(),
				entryOpacidadeTela.get(),
				entryFluorescenciaTela.get(),
				entryAlvuraD65Tela.get(),				
				entryLuminFeltro.get(),
				entryCoordAFeltro.get(),
				entryCoordBFeltro.get(),
				entryAlvuraISOFeltro.get(),
				entryBrancuraFeltro.get(),
				entryOpacidadeFeltro.get(),
				entryFluorescenciaFeltro.get(),
				entryAlvuraD65TelaFeltro.get()				
			]
			
			if "" not in entriesList:
				
				if definedRow == False:
					for i in range(2,1048576):
						# 1048576 is the max number of rows in excel
						c = resultSheet.cell(row=i, column=1)
						
						if c.value == None or c.value == "":
							emptyRow = i
							definedRow = True
							break

				date = str(datetime.datetime.now().date().strftime("%d-%m-%Y"))
				time = str(datetime.datetime.now().time())	

				resultSheet.cell(row=emptyRow, column=1).value = foNumber	
				resultSheet.cell(row=emptyRow, column=2).value = defect
				resultSheet.cell(row=emptyRow, column=3).value = register[:16]
				resultSheet.cell(row=emptyRow, column=4).value = comment
				resultSheet.cell(row=emptyRow, column=5).value = idString
				resultSheet.cell(row=emptyRow, column=6).value = analistNameEntry.get()
				resultSheet.cell(row=emptyRow, column=7).value = str(getpass.getuser())
				resultSheet.cell(row=emptyRow, column=8).value = date + " " + time[:8]
				resultSheet.cell(row=emptyRow, column=9).value = testType 
				resultSheet.cell(row=emptyRow, column=10).value = foTest
				resultSheet.cell(row=emptyRow, column=11).value = numberOfTestingPoints

				for n in range(len(entriesList)):
					columnPosition = n + 42
					resultSheet.cell(row=emptyRow, column=columnPosition).value = entriesList[n]
				
				try:
					wb.save("results.xlsx")
					messagebox.showinfo("Sucesso", "Resultados cadastrados com sucesso.")					
					
					analised.append(idString)

					other_index	= idValues.index(idString)

					idValues.remove(idString)

					del foNumberValues[other_index]
					del arriveDateValues[other_index]
					del arriveHourValues[other_index]
					del deliveringDateValues[other_index]
					del deliveringHourValues[other_index]
					del defectValues[other_index]
					del testTypeValues[other_index]
					del numberOfTestingPointsValues[other_index]
					del registerValues[other_index]
					del commentValues[other_index]
					del valuesList[other_index]
					
					definedRow = False

					print(f"Saved: valuesList length: {len(valuesList)}")
					# print(idValues)

					goBack()	

				except Exception as e:
					print(e)
					messagebox.showerror("Erro", "Não foi possível fazer o cadastro, a planilha está aberta em outro lugar.\nFeche-a e tente novamente.\n" + str(e))
							
			else:
			 	messagebox.showerror("Erro", "Preencha todos os campos.")			
		else:
			messagebox.showerror("Erro", "Digite seu Nome.")				

def goBack():
	global secondWindow
	global master_frame
	global foNumberBox

	if testType != "000-TESTES DE COR":
		for i in range(numberOfTestingPoints):
			testingPointsLabelList[i].destroy()
			testingPointsEntryList[i].destroy()

	elif testType == "000-TESTES DE COR":
		
		entryLuminTela.delete(0,END)
		entryCoordATela.delete(0,END)
		entryCoordBTela.delete(0,END)
		entryAlvuraISOTela.delete(0,END)
		entryOpacidadeTela.delete(0,END)
		entryFluorescenciaTela.delete(0,END)
		entryBrancuraTela.delete(0,END)
		entryAlvuraD65Tela.delete(0,END)

		entryLuminFeltro.delete(0,END)
		entryCoordAFeltro.delete(0,END)
		entryCoordBFeltro.delete(0,END)
		entryAlvuraISOFeltro.delete(0,END)
		entryBrancuraFeltro.delete(0,END)
		entryOpacidadeFeltro.delete(0,END)
		entryFluorescenciaFeltro.delete(0,END)
		entryAlvuraD65TelaFeltro.delete(0,END)

		entryLuminTela.grid_forget()
		entryCoordATela.grid_forget()
		entryCoordBTela.grid_forget()
		entryAlvuraISOTela.grid_forget()
		entryOpacidadeTela.grid_forget()
		entryFluorescenciaTela.grid_forget()
		entryBrancuraTela.grid_forget()
		entryAlvuraD65Tela.grid_forget()
		labelLuminTela.grid_forget()
		labelCoordATela.grid_forget()
		labelCoordBTela.grid_forget()
		labelAlvuraISOTela.grid_forget()
		labelOpacidadeTela.grid_forget()
		labelFluorescenciaTela.grid_forget()
		labelBrancuraTela.grid_forget()
		labelAlvuraD65Tela.grid_forget()

		entryLuminFeltro.grid_forget()
		entryCoordAFeltro.grid_forget()
		entryCoordBFeltro.grid_forget()
		entryAlvuraISOFeltro.grid_forget()
		entryBrancuraFeltro.grid_forget()
		entryOpacidadeFeltro.grid_forget()
		entryFluorescenciaFeltro.grid_forget()
		entryAlvuraD65TelaFeltro.grid_forget()
		labelLuminFeltro.grid_forget()
		labelCoordAFeltro.grid_forget()
		labelCoordBFeltro.grid_forget()
		labelAlvuraISOFeltro.grid_forget()
		labelBrancuraFeltro.grid_forget()
		labelOpacidadeFeltro.grid_forget()
		labelFluorescenciaFeltro.grid_forget()
		labelAlvuraD65TelaFeltro.grid_forget()

	testLabel['text'] = ""
	arriveDateLabel1["text"] = ""
	arriveHourLabel1["text"] = ""
	
	deliveringDateLabel1["text"] = ""
	deliveringDateLabel1["text"] = ""
	
	deliveringHourLabel1["text"] = ""	
	deliveringHourLabel1["text"] = ""
	
	defectLabel1["text"] = ""
	testTypeLabel1["text"] = ""
	numberOfTestingPointsLabel1["text"] = ""
	commentLabel1["text"] = ""
	
	foNumberBox.set("")

	analyseButton['state'] = DISABLED
		
	secondWindow.pack_forget()
	master_frame.pack()

def updateCombobox():	
	global registeredFos

	global comboBoxTextValues

	global registeredFoNumberValues
	global foNumberValues

	global registeredArriveDateValues
	global arriveDateValues
	
	global registeredArriveHourValues
	global arriveHourValues

	global registeredDeliveringDateValues
	global deliveringDateValues

	global registeredDeliveringHourValues
	global deliveringHourValues

	global registeredDefectValues
	global defectValues

	global registeredTestTypeValues
	global testTypeValues

	global registeredNumberOfTestingPointsValues
	global numberOfTestingPointsValues
	
	global registeredRegisterValues
	global registerValues

	global registeredCommentValues
	global commentValues

	global registeredValuesList
	global valuesList

	global registeredIdValues
	global idValues

	wb = load_workbook("db.xlsx")
	ws = wb["fosParaAnalise"]

	for z in range(2,1048576):
		# 1048576 is the max number of rows in excel
		a = ws.cell(row=z, column=1)
		b = ws.cell(row=z, column=2)
		c = ws.cell(row=z, column=3)
		d = ws.cell(row=z, column=4)
		e = ws.cell(row=z, column=5)
		f = ws.cell(row=z, column=6)
		g = ws.cell(row=z, column=7)
		h = ws.cell(row=z, column=8)
		I = ws.cell(row=z, column=9)
		j = ws.cell(row=z, column=10)
		k = ws.cell(row=z, column=11)
		l = ws.cell(row=z, column=12)
		
		if l.value != None and l.value != "":
			
			if l.value not in analisedFos and l.value not in idValues and l.value not in analised:
				registeredFos.append(a.value)
				comboBoxTextValues.append(a.value)

				registeredFoNumberValues.append(a.value)
				foNumberValues.append(a.value)

				registeredArriveDateValues.append(b.value)
				arriveDateValues.append(b.value)
				
				registeredArriveHourValues.append(c.value)
				arriveHourValues.append(c.value)

				registeredDeliveringDateValues.append(d.value)
				deliveringDateValues.append(d.value)

				registeredDeliveringHourValues.append(e.value)
				deliveringHourValues.append(e.value)

				registeredDefectValues.append(f.value)
				defectValues.append(f.value)

				registeredTestTypeValues.append(g.value)
				testTypeValues.append(g.value)

				registeredNumberOfTestingPointsValues.append(h.value)
				numberOfTestingPointsValues.append(h.value)
				
				registeredRegisterValues.append(I.value)
				registerValues.append(I.value)

				registeredCommentValues.append(j.value)
				commentValues.append(j.value)

				registeredValuesList.append(k.value)
				valuesList.append(k.value)

				registeredIdValues.append(l.value)
				idValues.append(l.value)

		elif l.value == None or l.value == "":
			print("Lists Rebuilt...")
			print(f"valuesList length: {len(idValues)}")
			break

	wb.close()

	testLabel['text'] = ""
	arriveDateLabel1["text"] = ""
	arriveHourLabel1["text"] = ""
	
	deliveringDateLabel1["text"] = ""
	deliveringDateLabel1["text"] = ""
	
	deliveringHourLabel1["text"] = ""	
	deliveringHourLabel1["text"] = ""
	
	defectLabel1["text"] = ""
	testTypeLabel1["text"] = ""
	numberOfTestingPointsLabel1["text"] = ""
	commentLabel1["text"] = ""
	
	# print("This is the Final Values List:", valuesList)
	foNumberBox["values"] = valuesList
	
def createTestResultEntriesWindow():
	global testingPointsLabelList
	global testingPointsEntryList

	master_frame.pack_forget()
	secondWindow.pack()			

	labelColumn = 0
	entryColumn = 1

	backButtonColumn = 0
	saveButtonColumn = 1
	buttonsColumnSpan = 1

	labelEntryRow = 0

	buttonsRow = 0

	firstLabel = Label(secondWindow, text="Insira seu nome e os resultados da Análise.")
	firstLabel.grid(row=0, column=0, columnspan=2, padx=xpadding, pady=15)

	analistNameLabel = Label(secondWindow, text="Insira seu Nome:")
	analistNameLabel.grid(row=1, column=0, padx=xpadding, pady=ypadding)

	analistNameEntry.grid(row=1, column=1, padx=xpadding, pady=ypadding)

	testLabel.grid(row=2, column=0, columnspan=2, sticky=W, padx=xpadding, pady=ypadding)
	testLabel['text'] = "Tipo de Analise: " + testType

	if testType == "000-TESTES DE COR":
		print("Teste de Cor...")
		
		labelLuminTela.grid(row=3, column=0, sticky=W, padx=xpadding, pady=ypadding)		
		entryLuminTela.grid(row=3, column=1, sticky=W, padx=xpadding, pady=ypadding)
		
		labelCoordATela.grid(row=4, column=0, sticky=W, padx=xpadding, pady=ypadding)		
		entryCoordATela.grid(row=4, column=1, sticky=W, padx=xpadding, pady=ypadding)
		
		labelCoordBTela.grid(row=5, column=0, sticky=W, padx=xpadding, pady=ypadding)		
		entryCoordBTela.grid(row=5, column=1, sticky=W, padx=xpadding, pady=ypadding)
		
		labelAlvuraISOTela.grid(row=6, column=0, sticky=W, padx=xpadding, pady=ypadding)		
		entryAlvuraISOTela.grid(row=6, column=1, sticky=W, padx=xpadding, pady=ypadding)
		
		labelBrancuraTela.grid(row=7, column=0, sticky=W, padx=xpadding, pady=ypadding)		
		entryBrancuraTela.grid(row=7, column=1, sticky=W, padx=xpadding, pady=ypadding)
		
		labelOpacidadeTela.grid(row=8, column=0, sticky=W, padx=xpadding, pady=ypadding)
		entryOpacidadeTela.grid(row=8, column=1, sticky=W, padx=xpadding, pady=ypadding)

		labelFluorescenciaTela.grid(row=9, column=0, sticky=W, padx=xpadding, pady=ypadding)		
		entryFluorescenciaTela.grid(row=9, column=1, sticky=W, padx=xpadding, pady=ypadding)		
		
		labelAlvuraD65Tela.grid(row=10, column=0, sticky=W, padx=xpadding, pady=ypadding)		
		entryAlvuraD65Tela.grid(row=10, column=1, sticky=W, padx=xpadding, pady=ypadding)

		labelLuminFeltro.grid(row=11, column=0, sticky=W, padx=xpadding, pady=ypadding)		
		entryLuminFeltro.grid(row=11, column=1, sticky=W, padx=xpadding, pady=ypadding)
		
		labelCoordAFeltro.grid(row=12, column=0, sticky=W, padx=xpadding, pady=ypadding)		
		entryCoordAFeltro.grid(row=12, column=1, sticky=W, padx=xpadding, pady=ypadding)
		
		labelCoordBFeltro.grid(row=13, column=0, sticky=W, padx=xpadding, pady=ypadding)		
		entryCoordBFeltro.grid(row=13, column=1, sticky=W, padx=xpadding, pady=ypadding)
		
		labelAlvuraISOFeltro.grid(row=14, column=0, sticky=W, padx=xpadding, pady=ypadding)		
		entryAlvuraISOFeltro.grid(row=14, column=1, sticky=W, padx=xpadding, pady=ypadding)
		
		labelBrancuraFeltro.grid(row=15, column=0, sticky=W, padx=xpadding, pady=ypadding)		
		entryBrancuraFeltro.grid(row=15, column=1, sticky=W, padx=xpadding, pady=ypadding)

		labelOpacidadeFeltro.grid(row=16, column=0, sticky=W, padx=xpadding, pady=ypadding)
		entryOpacidadeFeltro.grid(row=16, column=1, sticky=W, padx=xpadding, pady=ypadding)

		labelFluorescenciaFeltro.grid(row=17, column=0, sticky=W, padx=xpadding, pady=ypadding)		
		entryFluorescenciaFeltro.grid(row=17, column=1, sticky=W, padx=xpadding, pady=ypadding)		
		
		labelAlvuraD65TelaFeltro.grid(row=18, column=0, sticky=W, padx=xpadding, pady=ypadding)		
		entryAlvuraD65TelaFeltro.grid(row=18, column=1, sticky=W, padx=xpadding, pady=ypadding)		

		buttonsRow = 19

	else:	
		for index in range(numberOfTestingPoints):
			if index+1 > 10:
				labelColumn = 2
				entryColumn = 3
				
				buttonsColumnSpan = 2
				saveButtonColumn = 2

			if index+1 > 20:
				labelColumn = 4
				entryColumn = 5	

				buttonsColumnSpan = 3
				saveButtonColumn = 3

			if labelEntryRow >= 10:
				labelEntryRow = 0

			labelEntryRow += 1	

			buttonsRow += 5

			if index+1 >= 10:
				buttonsRow = 13		
			
			label = Label(secondWindow, text="P"+str(index+1)+":")
			testingPointsLabelList[index] = label
			testingPointsLabelList[index].grid(row=labelEntryRow+2, column=labelColumn, sticky=W, padx=xpadding, pady=ypadding)

			entry = Entry(secondWindow)
			testingPointsEntryList[index] = entry
			testingPointsEntryList[index].grid(row=labelEntryRow+2, column=entryColumn, sticky=W, padx=xpadding, pady=ypadding)

		if labelColumn == 0:
			labelColumn += 1
			entryColumn += 1

	backButton.grid(row=buttonsRow, column=backButtonColumn, columnspan=buttonsColumnSpan, padx=xpadding, pady=15)	
	
	saveResultsButton.grid(row=buttonsRow, column=saveButtonColumn, columnspan=buttonsColumnSpan, padx=xpadding, pady=15)
			
def getFoProperties(event):
	global foNumber
	global arriveDate
	global arriveHour
	global deliveringDate
	global deliveringHour
	global defect
	global testType
	global foTest
	global numberOfTestingPoints
	global register
	global comment
	global idString
	
	index = foNumberBox.current()

	if index != -1:
		foNumber = foNumberValues[index]
		arriveDate = arriveDateValues[index]
		arriveHour = arriveHourValues[index]
		deliveringDate = deliveringDateValues[index]
		deliveringHour = deliveringHourValues[index]
		defect = defectValues[index]
		testType = testTypeValues[index]
		foTest = valuesList[index]
		numberOfTestingPoints = numberOfTestingPointsValues[index]
		register = registerValues[index]
		comment = commentValues[index]

		idString = idValues[index]

		arriveDateLabel1["text"] = str(arriveDate)
		arriveHourLabel1["text"] = str(arriveHour)
		
		if str(deliveringDate) == "None":
			deliveringDateLabel1["text"] = ""
		else:		
			deliveringDateLabel1["text"] = str(deliveringDate)
		
		if str(deliveringHour) == "None":
			deliveringHourLabel1["text"] = ""	
		else:	
			deliveringHourLabel1["text"] = str(deliveringHour)
		
		defectLabel1["text"] = str(defect)
		testTypeLabel1["text"] = testType
		numberOfTestingPointsLabel1["text"] = str(numberOfTestingPoints)
		commentLabel1["text"] = comment

		analyseButton["state"] = NORMAL

	else:
		# print("Invalid Item on the Combobox")
		messagebox.showerror("Erro", "FO não existente.")

# Create all the widgets
master_frame = Frame(root)
master_frame.pack()

xpadding = 5
ypadding = 3

introLabel = Label(master_frame, text="Selecione a FO para Análise.")
introLabel.grid(row=0, column=0, columnspan=2, padx=xpadding, pady=15)

foNumberBoxLabel = Label(master_frame, text="Numero da FO:")
foNumberBoxLabel.grid(row=1, column=0, sticky=W, padx=xpadding, pady=ypadding)

foNumberBox = ttk.Combobox(master_frame, postcommand=updateCombobox, width=25)
foNumberBox.bind("<<ComboboxSelected>>", getFoProperties)
foNumberBox.grid(row=1, column=1, padx=xpadding, pady=ypadding)

arriveDateLabel = Label(master_frame, text="Data de Chegada:")
arriveDateLabel.grid(row=2, column=0, sticky=W, padx=xpadding, pady=ypadding)
arriveDateLabel1 = Label(master_frame)
arriveDateLabel1.grid(row=2, column=1, sticky=W, padx=xpadding, pady=ypadding)

arriveHourLabel = Label(master_frame, text="Hora de Chegada:")
arriveHourLabel.grid(row=3, column=0, sticky=W, padx=xpadding, pady=ypadding)
arriveHourLabel1 = Label(master_frame)
arriveHourLabel1.grid(row=3, column=1, sticky=W, padx=xpadding, pady=ypadding)

deliveringDateLabel = Label(master_frame, text="Data de Entrega:")
deliveringDateLabel.grid(row=4, column=0, sticky=W, padx=xpadding, pady=ypadding)
deliveringDateLabel1 = Label(master_frame)
deliveringDateLabel1.grid(row=4, column=1, sticky=W, padx=xpadding, pady=ypadding)

deliveringHourLabel = Label(master_frame, text="Hora da Entrega:")
deliveringHourLabel.grid(row=5, column=0, sticky=W, padx=xpadding, pady=ypadding)
deliveringHourLabel1 = Label(master_frame)
deliveringHourLabel1.grid(row=5, column=1, sticky=W, padx=xpadding, pady=ypadding)

defectLabel = Label(master_frame, text="Defeito:")
defectLabel.grid(row=6, column=0, sticky=W, padx=xpadding, pady=ypadding)
defectLabel1 = Label(master_frame)
defectLabel1.grid(row=6, column=1, sticky=W, padx=xpadding, pady=ypadding)

testTypeLabel = Label(master_frame, text="Tipo de Teste:")
testTypeLabel.grid(row=7, column=0, sticky=W, padx=xpadding, pady=ypadding)
testTypeLabel1 = Label(master_frame)
testTypeLabel1.grid(row=7, column=1, sticky=W, padx=xpadding, pady=ypadding)

numberOfTestingPointsLabel = Label(master_frame, text="Pontos de Analise:")
numberOfTestingPointsLabel.grid(row=8, column=0, sticky=W, padx=xpadding, pady=ypadding)
numberOfTestingPointsLabel1 = Label(master_frame)
numberOfTestingPointsLabel1.grid(row=8, column=1, sticky=W, padx=xpadding, pady=ypadding)

commentLabel = Label(master_frame, text="Observação:")
commentLabel.grid(row=9, column=0, sticky=NW, padx=xpadding, pady=ypadding)
commentLabel1 = Label(master_frame, anchor=W, wraplength=130, justify=LEFT)
commentLabel1.grid(row=9, column=1, sticky=W, padx=xpadding, pady=ypadding)

analyseButton = Button(master_frame, text="Analisar", command=createTestResultEntriesWindow, state=DISABLED, width=10)
analyseButton.grid(row=10, column=0, columnspan=2, padx=xpadding, pady=15)

# ///////////////////////////////////////////////////////////////////////////////////////////////////////
secondWindow = Frame(root)	

testLabel = Label(secondWindow)
	
analistNameEntry = Entry(secondWindow)

entryLuminTela = Entry(secondWindow)
entryCoordATela = Entry(secondWindow)
entryCoordBTela = Entry(secondWindow)
entryAlvuraISOTela = Entry(secondWindow)
entryBrancuraTela = Entry(secondWindow)
entryOpacidadeTela = Entry(secondWindow)
entryFluorescenciaTela = Entry(secondWindow)
entryAlvuraD65Tela = Entry(secondWindow)

labelLuminTela = Label(secondWindow, text="Lumin. L - TELA:")
labelCoordATela = Label(secondWindow, text="Coord. A - TELA:")
labelCoordBTela = Label(secondWindow, text="Coord. B - TELA:")
labelAlvuraISOTela = Label(secondWindow, text="Alvura ISO - TELA:")
labelBrancuraTela = Label(secondWindow, text="Brancura - TELA:")
labelOpacidadeTela = Label(secondWindow, text="Opacidade - TELA:")
labelFluorescenciaTela = Label(secondWindow, text="Fluorescência - TELA:")
labelAlvuraD65Tela = Label(secondWindow, text="Alvura D65 - TELA:")

entryLuminFeltro = Entry(secondWindow)
entryCoordAFeltro = Entry(secondWindow)
entryCoordBFeltro = Entry(secondWindow)
entryAlvuraISOFeltro = Entry(secondWindow)
entryBrancuraFeltro = Entry(secondWindow)
entryOpacidadeFeltro = Entry(secondWindow)
entryFluorescenciaFeltro = Entry(secondWindow)
entryAlvuraD65TelaFeltro = Entry(secondWindow)

labelLuminFeltro = Label(secondWindow, text="Lumin. L - FELTRO:")
labelCoordAFeltro = Label(secondWindow, text="Coord. A - FELTRO:")
labelCoordBFeltro = Label(secondWindow, text="Coord. B - FELTRO:")
labelAlvuraISOFeltro = Label(secondWindow, text="Alvura ISO - FELTRO:")
labelBrancuraFeltro = Label(secondWindow, text="Brancura - FELTRO:")
labelOpacidadeFeltro = Label(secondWindow, text="Opacidade - FELTRO:")
labelFluorescenciaFeltro = Label(secondWindow, text="Fluorescência - FELTRO:")
labelAlvuraD65TelaFeltro = Label(secondWindow, text="Alvura D65 - FELTRO:")
	
saveResultsButton = Button(secondWindow, text="Salvar", command=saveTestResults, width=10)
backButton = Button(secondWindow, text="Voltar", command=goBack, width=10)

root.mainloop()

# import os
# os.chdir("C:/Users/")
# os.getcwd()
# 'C:\\Users'

# list.remove(x), Remove the first item from the list whose value is x. It is an error if there is no such item.
