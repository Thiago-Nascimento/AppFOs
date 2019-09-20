from tkinter import * 
from tkinter import messagebox, ttk
from openpyxl import *
import datetime, os, sys, getpass

root = Tk()
root.title("Altera FO v1.0.0")
root.state("zoomed")

# os.chdir(r"\\polarsu2\Geqma\Engenharia da Qualidade\Laudos\FO\APP\BASE DADOS")

arriveDate = ""
arriveHour = ""
deliveringDate = ""
deliveringHour = ""
defect = ""
savingRow = None

arriveDateValues = []
registeredArriveDateValues = []

arriveHourValues = []
registeredArriveHourValues = []

deliveringDateValues= [] 
registeredDeliveringDateValues= [] 

deliveringHourValues = []
registeredDeliveringHourValues = []

defectValues = []
registeredDefectValues = []

emptyDeliveringDataFos = []
registeredEmptyDeliveringDataFos = []

savingRowList = []

def saveData():
	wb = load_workbook("db.xlsx")
	ws = wb["fosParaAnalise"]
	
	if deliveringDateEntry.get() == None or deliveringDateEntry.get() == "":
		messagebox.showerror("Erro", "Por Favor digite a data de entrega.")
		return

	if deliveringHourEntry.get() == None or deliveringHourEntry.get() == "":
		messagebox.showerror("Erro", "Por Favor digite o horário de entrega.")	
		return

	ws.cell(row=savingRow, column=4).value = deliveringDateEntry.get()
	ws.cell(row=savingRow, column=5).value = deliveringHourEntry.get()

	try:
		wb.save("db.xlsx")
		messagebox.showinfo("Sucesso", "Dados de Entrega alterados com sucesso.")

		del arriveDateValues[:]
		del registeredArriveDateValues[:]
		del arriveHourValues[:]
		del registeredArriveHourValues[:]
		del deliveringDateValues[:]
		del registeredDeliveringDateValues[:]
		del deliveringHourValues[:]
		del registeredDeliveringHourValues[:]
		del defectValues[:]
		del registeredDefectValues[:]
		del emptyDeliveringDataFos[:]
		del registeredEmptyDeliveringDataFos[:]
		del savingRowList[:]

		foNumberBox.delete(0,END)
		deliveringDateEntry.delete(0,END)
		deliveringHourEntry.delete(0,END)

		arriveDateLabel1["text"] = ""
		arriveHourLabel1["text"] = ""
		defectLabel1["text"] = ""

		changeButton["state"] = DISABLED

		updateCombobox()
	except Exception as e:
		print(e)
		messagebox.showerror("Erro", str(e))	

def getFoProperties(event):
	global arriveDate
	global arriveHour
	global defect
	global savingRow

	index = foNumberBox.current()

	if index != -1:
		arriveDate = arriveDateValues[index]
		arriveHour = arriveHourValues[index]
		defect = defectValues[index]
		savingRow = savingRowList[index]

		arriveDateLabel1["text"] = str(arriveDate)
		arriveHourLabel1["text"] = str(arriveHour)
		defectLabel1["text"] = str(defect)

		changeButton["state"] = NORMAL
	else:
		print("Inavlid Item in the Combobox")
		messagebox.showerror("Erro", "FO não existente.")	

def updateCombobox():
	global registeredArriveDateValues
	global arriveDateValues
	global registeredArriveHourValues
	global arriveHourValues
	global deliveringDateValues
	global registeredDeliveringDateValues
	global deliveringHourValues
	global registeredDeliveringHourValues
	global defectValues
	global registeredDefectValues
	global emptyDeliveringDataFos
	global registeredEmptyDeliveringDataFos
	global savingRowList
	
	wb = load_workbook("db.xlsx")
	ws = wb["fosParaAnalise"]

	for z in range(2,1048576):
		# 1048576 is the max number of rows in excel
		a = ws.cell(row=z, column=1)
		b = ws.cell(row=z, column=2)
		c = ws.cell(row=z, column=3)
		d = ws.cell(row=z, column=4)
		e = ws.cell(row=z, column=5)
		f = ws.cell(row=z, column=6)
		k = ws.cell(row=z, column=11)

		if a.value != None and a.value != "":

			if k.value not in emptyDeliveringDataFos:

				if d.value == None or d.value == "":
					arriveDateValues.append(b.value)
					registeredArriveDateValues.append(b.value)
					
					arriveHourValues.append(c.value)
					registeredArriveHourValues.append(c.value)
					
					deliveringDateValues.append(d.value)
					registeredDeliveringDateValues.append(d.value)

					deliveringHourValues.append(e.value)
					registeredDeliveringHourValues.append(e.value)

					defectValues.append(f.value)
					registeredDefectValues.append(f.value)
					
					emptyDeliveringDataFos.append(k.value)
					registeredEmptyDeliveringDataFos.append(k.value)

					savingRowList.append(z)

		elif a.value == None or a.value == "":
			print("Lists Built...")
			break

	wb.close()
	
	print("emptyDeliveringDataFos:",emptyDeliveringDataFos)
	foNumberBox["values"] = emptyDeliveringDataFos

xpadding = 5
ypadding = 3

mainframe = Frame(root)
mainframe.pack()

introLabel = Label(mainframe, text="Escolha a FO a ser alterada.")
introLabel.grid(row=0, column=0, columnspan=2, sticky=W, padx=xpadding, pady=15)

foNumberBoxLabel = Label(mainframe, text="Numero da FO:")
foNumberBoxLabel.grid(row=1, column=0, sticky=W, padx=xpadding, pady=ypadding)

foNumberBox = ttk.Combobox(mainframe, postcommand=updateCombobox)
foNumberBox.bind("<<ComboboxSelected>>", getFoProperties)
foNumberBox.grid(row=1, column=1, padx=xpadding, pady=ypadding, sticky=W)

arriveDateLabel = Label(mainframe, text="Data de Chegada:")
arriveDateLabel.grid(row=2, column=0, sticky=W, padx=xpadding, pady=ypadding)
arriveDateLabel1 = Label(mainframe)
arriveDateLabel1.grid(row=2, column=1, sticky=W, padx=xpadding, pady=ypadding)

arriveHourLabel = Label(mainframe, text="Hora de Chegada:")
arriveHourLabel.grid(row=3, column=0, sticky=W, padx=xpadding, pady=ypadding)
arriveHourLabel1 = Label(mainframe)
arriveHourLabel1.grid(row=3, column=1, sticky=W, padx=xpadding, pady=ypadding)

deliveringDateLabel = Label(mainframe, text="Data de Entrega:")
deliveringDateLabel.grid(row=4, column=0, sticky=W, padx=xpadding, pady=ypadding)
deliveringDateEntry = Entry(mainframe, width=23)
deliveringDateEntry.grid(row=4, column=1, sticky=W, padx=xpadding, pady=ypadding)

deliveringHourLabel = Label(mainframe, text="Hora da Entrega:")
deliveringHourLabel.grid(row=5, column=0, sticky=W, padx=xpadding, pady=ypadding)
deliveringHourEntry = Entry(mainframe, width=23)
deliveringHourEntry.grid(row=5, column=1, sticky=W, padx=xpadding, pady=ypadding)

defectLabel = Label(mainframe, text="Defeito:")
defectLabel.grid(row=6, column=0, sticky=W, padx=xpadding, pady=ypadding)
defectLabel1 = Label(mainframe)
defectLabel1.grid(row=6, column=1, sticky=W, padx=xpadding, pady=ypadding)

changeButton = Button(mainframe, text="Alterar", command=saveData, width=10, state=DISABLED)
changeButton.grid(row=7, column=0, columnspan=2, padx=xpadding, pady=15)

updateCombobox()

root.mainloop()
