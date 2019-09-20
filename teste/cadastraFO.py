from tkinter import *
from tkinter import messagebox, ttk
from openpyxl import *
import datetime, os

root = Tk()
root.title("Cadastra FO v1.0.0")

root.state('zoomed')

# Changing to the directory
os.chdir(r"\\polarsu2\Geqma\Engenharia da Qualidade\Laudos\FO\APP\BASE DADOS")

# Workbook and Worksheet loading
wb = load_workbook("db.xlsx")

print("Sheets:", wb.sheetnames)

ws = wb["fosParaAnalise"]
secondSheet = wb["fosSemAnalise"]
flawsDb = wb["defeitosDb"]

print("Main Sheet:", ws.title)
print("Defects Data Base:", flawsDb.title)

emptyRow = None
definedRow = False

foNumber = ""
arriveDate = ""
arriveHour = ""
deliveringDate = ""
deliveringHour = ""
defect = ""
numberOfTestingPoints = 0

isLabTest = IntVar()

flawList = []
typesList = []

for i in range(2, 1048576):		# This gets the defects from a list
	# 1048576 is the max number of rows in excel
	flaw = flawsDb.cell(row=i, column=1)

	if flaw.value != None and flaw.value != "":
		flawList.append(flaw.value)

	elif flaw.value == None or flaw.value == "":
		print("Defects list loaded...")
		# print("Defects List:", flawList)
		break

for j in range(2, 1048576):		# this gets the tests from a list
	# 1048576 is the max number of rows in excel
	test = flawsDb.cell(row=j, column=2)

	if test.value != None and test.value != "":
		typesList.append(test.value)

	elif test.value == None or test.value == "":
		print("Test list loaded...")
		# print("Test List:", typesList)
		break

def updateDb():
	global foNumber
	global arriveDate
	global arriveHour
	global deliveringDate
	global deliveringHour
	global defect
	global testType
	global numberOfTestingPoints
	global comment
	global definedRow
	global emptyRow

	if definedRow == False:
		for i in range(2,1048576):
			# 1048576 is the max number of rows in excel
			
			if isLabTest.get() == 1:
				c = ws.cell(row=i, column=1)

			elif isLabTest.get() == 0:	
				c = secondSheet.cell(row=i, column=1)

			if c.value == None or c.value == "":
				emptyRow = i
				definedRow = True
				break

	print('Saving Row:', emptyRow)

	date = str(datetime.datetime.now().date().strftime("%d-%m-%Y"))
	time = str(datetime.datetime.now().time())

	if isLabTest.get() == 1:	
		ws.cell(row=emptyRow, column=1).value = int(foNumber)
		ws.cell(row=emptyRow, column=2).value = arriveDate
		ws.cell(row=emptyRow, column=3).value = arriveHour
		ws.cell(row=emptyRow, column=4).value = deliveringDate
		ws.cell(row=emptyRow, column=5).value = deliveringHour
		ws.cell(row=emptyRow, column=6).value = defect
		ws.cell(row=emptyRow, column=7).value = testType
		ws.cell(row=emptyRow, column=8).value = int(numberOfTestingPoints)
		ws.cell(row=emptyRow, column=9).value = date + " " + time[:8]
		ws.cell(row=emptyRow, column=10).value = comment
		ws.cell(row=emptyRow, column=11).value = str(foNumber) + " - " + str(testType[4:])
		ws.cell(row=emptyRow, column=12).value = date + " " + time[:8] + " - " + comment

	elif isLabTest.get() == 0:
		secondSheet.cell(row=emptyRow, column=1).value = int(foNumber)
		secondSheet.cell(row=emptyRow, column=2).value = arriveDate
		secondSheet.cell(row=emptyRow, column=3).value = arriveHour
		secondSheet.cell(row=emptyRow, column=4).value = deliveringDate
		secondSheet.cell(row=emptyRow, column=5).value = deliveringHour
		secondSheet.cell(row=emptyRow, column=6).value = defect		
		secondSheet.cell(row=emptyRow, column=7).value = date + " " + time[:8]
		secondSheet.cell(row=emptyRow, column=8).value = comment		
		secondSheet.cell(row=emptyRow, column=9).value = str(foNumber) + " - " + str(testType[4:])

	try:
		wb.save("db.xlsx")
		print("Successfully Saved!")
		messagebox.showinfo("Sucesso", "Análise cadastrada com sucesso.")

		definedRow = False

		# clearEntries()
	except:
		print("ERROR: WorkBook is Open Somewhere Else...")		
		messagebox.showerror("Erro", "Não foi possível fazer o cadastro, a planilha está aberta em outro lugar.\nFeche-a e tente novamente.")
		
def clearEntries():
	# Clear the entries
	global definedRow
	global isLabTest
	
	definedRow = False

	foNumberEntry.delete(0, END)
	arriveDateEntry.delete(0, END)
	arriveHourEntry.delete(0, END)
	deliveringDateEntry.delete(0, END)
	deliveringHourEntry.delete(0, END)
	defectBox.delete(0, END)
	testTypeBox.delete(0, END)
	numberOfTestingPointsEntry.delete(0, END)
	commentEntry.delete(0, END)

	print(isLabTest.get())

def registerFo():
	# Validate all the entries and the combobox
	global foNumber
	global arriveDate
	global arriveHour
	global deliveringDate
	global deliveringHour
	global defect
	global testType
	global numberOfTestingPoints
	global comment

	if foNumberEntry.get() == None or foNumberEntry.get() == "":
		messagebox.showerror("Erro", "Por favor digite o número da FO.")
		return

	if arriveDateEntry.get() == None or arriveDateEntry.get() == "":
		messagebox.showerror("Erro", "Por favor digite a data de chegada da FO.")
		return

	if arriveHourEntry.get() == None or arriveHourEntry.get() == "":
		messagebox.showerror("Erro", "Por favor digite a hora de chegada da FO.")		
		return

	if defectBox.get() == None or defectBox.get() == "":
			messagebox.showerror("Erro", "Por favor selecione o defeito.")
			return

	if isLabTest == 1:
		if testTypeBox.get() == None or testTypeBox.get() == "":
			messagebox.showerror("Erro", "Por favor selecione o tipo de análise.")	
			return
		
		if numberOfTestingPointsEntry.get() == None or numberOfTestingPointsEntry.get() == "":
			messagebox.showerror("Erro", "Por favor especifique o número de pontos.")
			return

	foNumber = foNumberEntry.get()
	arriveDate = arriveDateEntry.get()
	arriveHour = arriveHourEntry.get()
	deliveringDate = deliveringDateEntry.get()
	deliveringHour = deliveringHourEntry.get()
	defect = defectBox.get()
	testType = testTypeBox.get()
	numberOfTestingPoints = numberOfTestingPointsEntry.get()
	comment = commentEntry.get()

	print("foNumber:", foNumber)
	print("arriveDate:", arriveDate)
	print("arriveHour:", arriveHour)
	print("deliveringDate:", deliveringDate)
	print("deliveringHour:", deliveringHour)
	print("defect:", defect)
	print("testType:", testType)
	print("numberOfTestingPoints:", numberOfTestingPoints)
	print("comment:", comment)

	updateDb()

# Create all the widgets
xpadding = 5
ypadding = 3

mainframe = Frame(root)
mainframe.pack()

introLabel = Label(mainframe, text="Insira as Informações.")
introLabel.grid(row=0, column=0, columnspan=2, padx=xpadding, pady=15)

foNumberEntryLabel = Label(mainframe, text="Numero da FO:")
foNumberEntryLabel.grid(row=1, column=0, sticky=W, padx=xpadding, pady=ypadding)

foNumberEntry = Entry(mainframe, width=40)
foNumberEntry.grid(row=1, column=1, padx=xpadding, pady=ypadding)

arriveDateEntryLabel = Label(mainframe, text="Data de Chegada:")
arriveDateEntryLabel.grid(row=2, column=0, sticky=W, padx=xpadding, pady=ypadding)

arriveDateEntry = Entry(mainframe, width=40)
arriveDateEntry.grid(row=2, column=1, padx=xpadding, pady=ypadding)

arriveHourEntryLabel = Label(mainframe, text="Horario de Chegada:")
arriveHourEntryLabel.grid(row=3,column=0, sticky=W, padx=xpadding, pady=ypadding)

arriveHourEntry = Entry(mainframe, width=40)
arriveHourEntry.grid(row=3, column=1, padx=xpadding, pady=ypadding)

deliveringDateEntryLabel = Label(mainframe, text="Data de Entrega:")
deliveringDateEntryLabel.grid(row=4, column=0, sticky=W, padx=xpadding, pady=ypadding)

deliveringDateEntry = Entry(mainframe, width=40)
deliveringDateEntry.grid(row=4, column=1, padx=xpadding, pady=ypadding)

deliveringHourEntryLabel = Label(mainframe, text="Horario de Entrega:")
deliveringHourEntryLabel.grid(row=5, column=0, sticky=W, padx=xpadding, pady=ypadding)

deliveringHourEntry = Entry(mainframe, width=40)
deliveringHourEntry.grid(row=5, column=1, padx=xpadding, pady=ypadding)

defectBoxLabel = Label(mainframe, text="Defeito:")
defectBoxLabel.grid(row=6, column=0, sticky=W, padx=xpadding, pady=ypadding)

defectBox = ttk.Combobox(mainframe, values=flawList, width=37)
defectBox.grid(row=6, column=1, padx=xpadding, pady=ypadding)

testTypeBoxLabel = Label(mainframe, text="Tipo de Teste:")
testTypeBoxLabel.grid(row=7, column=0, sticky=W, padx=xpadding, pady=ypadding)

testTypeBox = ttk.Combobox(mainframe, values=typesList, width=37)
testTypeBox.grid(row=7, column=1, padx=xpadding, pady=ypadding)

numberOfTestingPointsLabel = Label(mainframe, text="Número de Pontos a Analisar:")
numberOfTestingPointsLabel.grid(row=8, column=0, sticky=W, padx=xpadding, pady=ypadding)

numberOfTestingPointsEntry = Entry(mainframe, width=40)
numberOfTestingPointsEntry.grid(row=8, column=1, padx=xpadding, pady=ypadding)

labTestCheckbutton = Checkbutton(mainframe, text="Analisar no Lab", variable=isLabTest, onvalue=1, offvalue=0)
labTestCheckbutton.grid(row=9, column=0, padx=xpadding, pady=ypadding, sticky=W)

commentLabel = Label(mainframe, text="Observações:")
commentLabel.grid(row=10, column=0, sticky=W, padx=xpadding, pady=15)

commentEntry = Entry(mainframe, width=40)
commentEntry.grid(row=10, column=1, padx=xpadding, pady=15)

clearEntriesButton = Button(mainframe, text="Limpar", command=clearEntries, width=10)
clearEntriesButton.grid(row=11, column=0, sticky=E, padx=xpadding, pady=15)

registerFoButton = Button(mainframe, text="Cadastrar", command=registerFo, width=10)
registerFoButton.grid(row=11, column=1, sticky=W, padx=xpadding, pady=15)

# print(datetime.datetime.now().strftime("%d-%m-%y-%H-%M"))
		
root.mainloop()

# import os
# os.chdir("C:/Users/")
# os.getcwd()
# 'C:\\Users'

# pyinstaller --onefile --noconsole cadastraFOv2.py
